import io
from django.contrib.auth.models import User
from django.http import HttpResponse
from rest_framework.views import APIView
from .permissions import IsStaff
from .models import SessionPackage, AttendanceRecord


# ── helpers ──────────────────────────────────────────────────────────────────

def _patient_report_data(patient: User) -> dict:
    packages = SessionPackage.objects.filter(patient=patient).prefetch_related(
        "attendance_records"
    ).select_related("plan").order_by("-purchased_at")

    pkg_rows = []
    for pkg in packages:
        records = pkg.attendance_records.order_by("date")
        pkg_rows.append({
            "pkg": pkg,
            "records": list(records),
        })

    return {"patient": patient, "packages": pkg_rows}


def _all_patients_data() -> list:
    patients = User.objects.filter(is_staff=False, is_active=True).order_by("last_name", "first_name")
    rows = []
    for p in patients:
        packages = SessionPackage.objects.filter(patient=p).select_related("plan").order_by("-is_active", "-purchased_at")
        if packages.exists():
            for pkg in packages:
                rows.append({
                    "patient": p,
                    "pkg": pkg,
                    "came": pkg.attendance_records.filter(status="came").count(),
                    "no_show": pkg.attendance_records.filter(status="no_show").count(),
                })
        else:
            rows.append({
                "patient": p,
                "pkg": None,
                "came": 0,
                "no_show": 0,
            })
    return rows


# ── Excel builders ────────────────────────────────────────────────────────────

def _build_patient_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    patient = data["patient"]
    wb = Workbook()

    # Sheet 1: Paket Özeti
    ws1 = wb.active
    ws1.title = "Paket Özeti"
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")

    def write_header(ws, cols, row=1):
        for col, title in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    ws1.append([f"Öğrenci: {patient.get_full_name() or patient.username}  |  E-posta: {patient.email}"])
    ws1.append([])
    pkg_headers = ["Paket Adı", "Toplam Seans", "Geldi", "Gelmedi", "Kalan", "Fiyat", "Ödendi", "Başlangıç", "Durum"]
    write_header(ws1, pkg_headers, row=3)

    for item in data["packages"]:
        pkg = item["pkg"]
        ws1.append([
            pkg.name or (pkg.plan.name if pkg.plan else "—"),
            pkg.total_sessions,
            pkg.used_sessions,
            pkg.no_show_count,
            pkg.remaining_sessions,
            float(pkg.price) if pkg.price else 0,
            "Evet" if pkg.is_paid else "Hayır",
            str(pkg.purchased_at),
            "Aktif" if pkg.is_active else "Pasif",
        ])

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    # Sheet 2: Devam Geçmişi
    ws2 = wb.create_sheet("Devam Geçmişi")
    att_headers = ["Tarih", "Durum", "Paket", "Not"]
    write_header(ws2, att_headers)

    for item in data["packages"]:
        pkg = item["pkg"]
        pkg_name = pkg.name or (pkg.plan.name if pkg.plan else "—")
        for rec in item["records"]:
            ws2.append([
                str(rec.date),
                "Geldi" if rec.status == "came" else "Gelmedi",
                pkg_name,
                rec.note or "",
            ])

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_all_xlsx(rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Öğrenci Raporu"
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Ad Soyad", "E-posta", "Paket", "Toplam Seans", "Geldi", "Gelmedi", "Kalan Seans", "Ödeme", "Durum"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        p = r["patient"]
        pkg = r["pkg"]
        ws.append([
            p.get_full_name() or p.username,
            p.email,
            (pkg.name or (pkg.plan.name if pkg.plan else "—")) if pkg else "—",
            pkg.total_sessions if pkg else 0,
            r["came"],
            r["no_show"],
            pkg.remaining_sessions if pkg else 0,
            ("Ödendi" if pkg.is_paid else "Bekliyor") if pkg else "—",
            ("Aktif" if pkg.is_active else "Pasif") if pkg else "—",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF builders ──────────────────────────────────────────────────────────────

def _build_patient_pdf(data: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    patient = data["patient"]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    blue = colors.HexColor("#1D4ED8")
    story = []

    title_style = ParagraphStyle("title", parent=styles["Heading1"], textColor=blue, fontSize=16)
    story.append(Paragraph(f"Öğrenci Raporu: {patient.get_full_name() or patient.username}", title_style))
    story.append(Paragraph(f"E-posta: {patient.email}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    for item in data["packages"]:
        pkg = item["pkg"]
        pkg_name = pkg.name or (pkg.plan.name if pkg.plan else "—")
        story.append(Paragraph(f"<b>{pkg_name}</b> — {'Aktif' if pkg.is_active else 'Pasif'}", styles["Heading2"]))

        summary = [
            ["Toplam", "Geldi", "Gelmedi", "Kalan", "Ödeme"],
            [pkg.total_sessions, pkg.used_sessions, pkg.no_show_count, pkg.remaining_sessions,
             "Ödendi" if pkg.is_paid else "Bekliyor"],
        ]
        t = Table(summary, colWidths=[3*cm]*5)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), blue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

        if item["records"]:
            att_data = [["Tarih", "Durum", "Not"]]
            for rec in item["records"]:
                att_data.append([
                    str(rec.date),
                    "Geldi" if rec.status == "came" else "Gelmedi",
                    rec.note or "",
                ])
            at = Table(att_data, colWidths=[4*cm, 3*cm, 9*cm])
            at.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B82F6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            story.append(at)

        story.append(Spacer(1, 0.6*cm))

    doc.build(story)
    return buf.getvalue()


def _build_all_pdf(rows: list) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    blue = colors.HexColor("#1D4ED8")
    story = []

    story.append(Paragraph("Tüm Öğrenciler Raporu", ParagraphStyle("t", parent=styles["Heading1"], textColor=blue, fontSize=16)))
    story.append(Spacer(1, 0.4*cm))

    data = [["Ad Soyad", "E-posta", "Paket", "Toplam", "Geldi", "Gelmedi", "Kalan", "Ödeme", "Durum"]]
    for r in rows:
        p = r["patient"]
        pkg = r["pkg"]
        data.append([
            p.get_full_name() or p.username,
            p.email,
            (pkg.name or (pkg.plan.name if pkg.plan else "—")) if pkg else "—",
            pkg.total_sessions if pkg else 0,
            r["came"],
            r["no_show"],
            pkg.remaining_sessions if pkg else 0,
            ("Ödendi" if pkg.is_paid else "Bekliyor") if pkg else "—",
            ("Aktif" if pkg.is_active else "Pasif") if pkg else "—",
        ])

    col_w = [3.5*cm, 4.5*cm, 3.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2.5*cm, 2*cm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


# ── Views ─────────────────────────────────────────────────────────────────────

class AdminPatientReportView(APIView):
    permission_classes = [IsStaff]

    def get(self, request, pk):
        patient = User.objects.filter(pk=pk, is_staff=False).first()
        if not patient:
            return HttpResponse(status=404)

        fmt = request.query_params.get("report_format", "xlsx")
        data = _patient_report_data(patient)
        name = (patient.get_full_name() or patient.username).replace(" ", "_")

        if fmt == "pdf":
            content = _build_patient_pdf(data)
            response = HttpResponse(content, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{name}_rapor.pdf"'
        else:
            content = _build_patient_xlsx(data)
            response = HttpResponse(content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename="{name}_rapor.xlsx"'

        return response


class AdminAllPatientsReportView(APIView):
    permission_classes = [IsStaff]

    def get(self, request):
        fmt = request.query_params.get("report_format", "xlsx")
        rows = _all_patients_data()

        if fmt == "pdf":
            content = _build_all_pdf(rows)
            response = HttpResponse(content, content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="ogrenciler_raporu.pdf"'
        else:
            content = _build_all_xlsx(rows)
            response = HttpResponse(content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="ogrenciler_raporu.xlsx"'

        return response
